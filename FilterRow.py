class FilterRow:
    
    def __init__(self, ctrl_path, ctrl_sheet):
        self.ctrl_path = ctrl_path
        self.ctrl_sheet = ctrl_sheet
        self.controlsheet = None
        
        have_upenpyxl = False
        try:
            from openpyxl import load_workbook
            have_upenpyxl = True
        except ModuleNotFoundError:
            print("openpyxl not found!")
            print("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has openpyxl installed.")

        if have_upenpyxl:
            try:
                controlbook  = load_workbook(self.ctrl_path)
                if self.ctrl_sheet in controlbook.sheetnames:
                    self.controlsheet = controlbook[self.ctrl_sheet]
                    print(f"Reading \"{self.ctrl_sheet}\" worksheet in {self.ctrl_path}")
                else :
                    print(f"Did not locate \"{self.ctrl_sheet}\" worksheet in {self.ctrl_path}")

            except FileNotFoundError:
                print(f"File {self.ctrl_path} not found.")
            except Exception as err:
                print(f"Could not open the file: {err}")

                