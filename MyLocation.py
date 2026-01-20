# MyLocation.py - Load "My Location" value into worksheet - based on "Project Location", and using a location worksheet as input.

# "My Location" will be used in conjunction with "Co-Location" to filter out Roles for which the Co-Location requirement is untenable.

# Suggested Enhancements:
#   Add input worksheet for client locations.
#   Provide graceful exit if worksheet is open in Excel session.
#   Add input re unidentified locations.  Ask if want to add from command line - or directly into spreadsheet.

import sys
import shlex
from pathlib import Path
from datetime import datetime
import ReadControl

arguments = sys.argv[0:]  # List of all arguments.  arguments[0] = <script name>  arguments[1] = <Control File name>   arguments[2] = <Input worksheet File name>
now_time = datetime.now()
start_message = "Python script " + arguments[0] + " started at " + now_time.strftime("%Y-%m-%d %H:%M:%S")

if len(arguments) < 3 :
    print("Missing required command line parameters.")
    print(arguments)
    print(f"Usage: python {sys.argv[0]} <control file> <workbook>")
    print(" .. or, if using BAT file: MyLocation.bat <workbook>")
    print("Terminating process.")
    sys.exit(1)

worksheet_path = Path(arguments[2])
if not worksheet_path.is_file():
    print(f"Input worksheet \"{arguments[2]}\" not found.")
    print("Terminating process.")
    sys.exit(1)

# Read parameter file and load values into a dictionary object.
err_code, parms_dict = ReadControl.read(arguments[1], True)
if err_code > 0 :
    print("Terminating process.")
    sys.exit(1)

# MessageOut(), MessageClose(), and MessageShow() functions for output message handling. [Development note: Should add code to handle exceptions.]
mssg_file = None
mssg_file_path = None
def MessageOut(mssg_txt) :  # Write message to output message file.
    global mssg_file, mssg_file_path
    if not mssg_file :
       mssg_file = open(mssg_file_path, "w")
    mssg_file.write(f"{mssg_txt}\n")
def MessageClose() :
    global mssg_file, mssg_file_path
    if mssg_file :
        MessageOut(" ")
        MessageShow("Output messages to: " + mssg_file_path)
        mssg_file.close()
def MessageShow(mssg_txt) :  # Display message to console and write to output message file.
    print(mssg_txt)
    MessageOut(mssg_txt)
    
# Write start message to output message file.
mssg_file_path = parms_dict['messagdir'][0] + "\\MyLocation.mssg"  # Message Directory obtained from Parameter File.
MessageOut(start_message)
MessageOut(" ")

# print("Python executable: ", sys.executable)
try:
    import pandas as pd
#   print("Pandas location:", pd.__file__)
#   print("Pandas version:", pd.__version__)
except ModuleNotFoundError:
    MessageShow("Pandas not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has Pandas installed.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    MessageShow("openpyxl not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has openpyxl installed.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

if "col_projlocat" in parms_dict:
    ok = True
else :
    MessageShow(f"Missing \"col_projlocat\" parameter from {arguments[1]} Control File.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

# Read Location Worksheet and load values into a dictionary object.
control_path = Path(parms_dict['locatsheet'][0])   # The Location Worksheet path is specified in the Control File.
if not control_path.is_file():
    MessageShow(f"Location Worksheet \"{parms_dict['locatsheet'][0]}\" not found.")
    MessageShow(f"Make sure that Control File {arguments[1]} specifies the full worksheet path correctly.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

try:
    locbook  = load_workbook(parms_dict['locatsheet'][0])
    locsheet = locbook.active  # Get the active (only) worksheet.
except FileNotFoundError:
    MessageShow(f"File {parms_dict['locatsheet'][0]} not found.")
except Exception as err:
    MessageShow(f"Could not open the file: {err}")

locats_dict = {}
for row in locsheet.iter_rows(min_row=2, max_row=locsheet.max_row, values_only=True):
    if row[0] is not None :
        locats_dict[row[0]] = row[1]

MessageShow(f"Read {len(locats_dict)} locations from {parms_dict['locatsheet'][0]}.")



#   Build dictionary from parameters for comparing which value to return for a mulit-value Project Location input.  (The result should refer to the closest Project Location.)
loc_compare_dict = {}
loc_compare_max  = 0
def loadLocCampare(compare_list) :
    global arguments, parms_dict, loc_compare_dict, loc_compare_max
    
    parms_error = False
    for compname in compare_list :
        try:
            comp_list = parms_dict[compname]
            loc_rating_s = comp_list[0]
            try :
                loc_rating = int(loc_rating_s)
            except ValueError :
                MessageShow(f"First element in \"{compname}\" parameter list is \"{loc_rating_s}\" but was expecting an integer value.")
                loc_rating = loc_compare_max
                parms_error = True
           
            if loc_rating > loc_compare_max :
                loc_compare_max = loc_rating
                
            comp_key = comp_list[1]
            loc_compare_dict[comp_key] = loc_rating
    #   Development note: At this point, only processing up to three elements in the list: <loc_rating> <comp_key_1> <comp_key_2>
            if len(comp_list) > 2 :
                comp_key = comp_list[2]
                loc_compare_dict[comp_key] = loc_rating
        except KeyError:
            MessageShow(f"\"{compname}\" parameter listed in \"locompare\" but is not matched in the {arguments[1]} Control File.")
            parms_error = True
    
    if parms_error :
        return False
    return True


# Read Roles Worksheet and process.
try:
    rolebook  = load_workbook(arguments[2])
    
    sheet_to_remove = "Instructions"    # Drop the "Instructions" worksheet.
    # Note: For some reason, this script causes the "Instructions" worksheet to get blanked out.  Have not figured out how to keep this from happening - so just deleting the worksheet.
    if sheet_to_remove in rolebook.sheetnames:
        drop_sheet = rolebook[sheet_to_remove]
        rolebook.remove(drop_sheet)
        MessageShow(f"Deleting worksheet '{sheet_to_remove}'.")
        
    rolesheet_title = rolebook.sheetnames[0]
    rolesheet = rolebook.active
    MessageShow(f"Processing \"{rolesheet_title}\" worksheet.")
    
except FileNotFoundError:
    MessageShow(f"File {arguments[2]} not found.")
except InvalidFileException:
    MessageShow(f"The file '{arguments[2]}' is not a valid Excel file or is corrupted.")
except PermissionError:
    MessageShow(f"Permission denied. Please close '{arguments[2]}' if it is open in Excel.")
except Exception as unexpected_error:
    print(f"An unexpected error occurred: {unexpected_error}")


role_headers = {}
def getRoleHeaders (sheet_row) :
    global start_data_row, role_headers
    for idx, cell in enumerate(sheet_row):
        if cell.value is not None :
            col_header = cell.value.strip() 
            role_headers[col_header] = idx
            
start_data_row = 2
getRoleHeaders(rolesheet[1])
if "skip_header" in parms_dict :    # Special processing for GPS_Open_Demands_Report.xlsx - headers needed for processing are on second line.
    skip_header = parms_dict['skip_header'][0]
    if skip_header in role_headers and role_headers[skip_header] == 0 :
        start_data_row += 1 
        getRoleHeaders(rolesheet[2])

have_col_headers = True
try:
    if parms_dict['col_request'][0] in role_headers:
        ix_request = role_headers[parms_dict['col_request'][0]]
        MessageOut(f"ix_request = {ix_request}")
    else:
        MessageShow(f"Did not locate \"{parms_dict['col_request'][0]}\" column header in {arguments[2]} workbook.")
        have_col_headers = False
except KeyError:
    MessageShow(f"\"col_request\" parameter is not specified in the {arguments[1]} Control File.")
    have_col_headers = False

try:   
    if parms_dict['col_projlocat'][0] in role_headers:
        ix_proj_locat = role_headers[parms_dict['col_projlocat'][0]]
        MessageOut(f"ix_proj_locat = {ix_proj_locat}")
    else:
        MessageShow(f"Did not locate \"{parms_dict['col_projlocat'][0]}\" column header in {arguments[2]} workbook.")
        have_col_headers = False
except KeyError:
    MessageShow(f"\"col_projlocat\" parameter is not specified in the {arguments[1]} Control File.")
    have_col_headers = False

location_col_inserted = False
try:
    if parms_dict['col_location'][0] in role_headers:
        ix_my_locat = role_headers[parms_dict['col_location'][0]]
        MessageOut(f"ix_my_locat = {ix_my_locat}")
    else:
        if have_col_headers :
            rolesheet.insert_cols(ix_proj_locat + 2)    # Column location count from A = 1
            ix_my_locat = ix_proj_locat + 1             # Column location count from A = 0
            MessageShow(f"Inserting \"{parms_dict['col_location'][0]}\" column in {arguments[2]} workbook.")
            MessageOut(f"ix_my_locat = {ix_my_locat}")
            rolesheet.cell(row=start_data_row - 1, column=ix_my_locat + 1, value=parms_dict['col_location'][0])      # Column location count from A = 1
            location_col_inserted = True
        else:
            MessageShow(f"Did not locate \"{parms_dict['col_location'][0]}\" column header in {arguments[2]} workbook.")
except KeyError:
    MessageShow(f"\"col_location\" parameter is not specified in the {arguments[1]} Control File.")
    have_col_headers = False

ix_req_offc = -1
try:
    if parms_dict['col_reqoffice'][0] in role_headers:
        ix_req_offc = role_headers[parms_dict['col_reqoffice'][0]]
        MessageOut(f"ix_req_offc = {ix_req_offc}")
    else:
        MessageOut(f"Did not locate (optional) \"{parms_dict['col_reqoffice'][0]}\" column header in {arguments[2]} workbook.")
        # Not always required for processing - so not stopping process.
 
except KeyError:
    MessageShow(f"(Optional) \"col_reqoffice\" parameter is not specified in the {arguments[1]} Control File.")
    
# Build dictionary from parameters for comparing which value to return for a mulit-value Project Location input. 
try:
    compare_list = parms_dict['locompare']
    if not loadLocCampare(compare_list) :
        have_col_headers = False 
except KeyError:
    MessageShow(f"\"locompare\" parameter is not specified in the {arguments[1]} Control File.")
    have_col_headers = False

if not have_col_headers :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

if location_col_inserted :
    if ix_request > ix_proj_locat :
        ix_request += 1
    if ix_req_offc > ix_proj_locat :
        ix_req_offc += 1
  

loc_unknown_list = []   # For collecting unique unknown locations

#   processLocation() function
def processLocation(proj_location) :
    global locats_dict, count_matched
    return_location = ""
    
    if proj_location in locats_dict:
        return_location = locats_dict[proj_location]
        count_matched += 1
    else :
        return_location = "Unknown"
        if proj_location not in loc_unknown_list:
            loc_unknown_list.append(proj_location)
    return return_location
    

#   processLocations() function - test if Project Location contains multiple values (delimited by "|" bar character).
def processLocations(proj_location) :
    global count_multi, count_matched
    return_location = ""
    
    if proj_location is None or proj_location == "" :
        return_location = "Non"
        count_matched += 1
        return return_location
        
    if isinstance(proj_location, int) :
        return_location = processLocation(proj_location)
        return return_location
    
    split_line = proj_location.split('|')
    if len(split_line) > 1 :
        count_multi += len(split_line) - 1
        compare_return = "xxx"
        compare_rank   = loc_compare_max + 1
        for split_value in split_line :
            if split_value != "" :
                test_location = processLocation(split_value)
                if test_location in loc_compare_dict :
                    test_rank = loc_compare_dict[test_location]
                else :
                    test_rank = loc_compare_max + 1

                if test_rank < compare_rank :
                    compare_return = test_location
                    compare_rank   = test_rank
        
        return_location = compare_return
    else :
        return_location = processLocation(proj_location)
    return return_location
    
 
count_matched = 0
count_rows = 0
count_multi = 0
for rrow in rolesheet.iter_rows(min_row=start_data_row) :
    proj_location = rrow[ix_proj_locat].value
    request_id    = rrow[ix_request].value
    
    if request_id is not None and request_id != "" :
        count_rows += 1
         
        if ix_req_offc > -1 and proj_location == "Deloitte Office" :     # Special processing for MySource worksheet
            proj_location = rrow[ix_req_offc].value
    
   
        update_cell = rrow[ix_my_locat]
        update_cell.value = processLocations(proj_location)

rolebook.save(arguments[2])

multi_message_string = "."
if count_multi > 0 :
    multi_message_string = f" + {count_multi} additional (multi) values = {count_rows + count_multi}."
MessageShow(f"Matched Project Location for {count_matched} out of {count_rows} rows" + multi_message_string)
    
MessageShow(f"{len(loc_unknown_list)} unique Project Locations still not identified.")
for loc_unk in loc_unknown_list :
    MessageOut(loc_unk)
MessageClose()