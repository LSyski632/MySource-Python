# MySource.py - Filter Open Roles spreadsheet and drop Roles that are not suitable.  Match Roles by Request # against Action File to identify Roles previously reviewed (Applied for or Passed).

# Suggested Enhancements:
#       Add logic to catch None in colocation.  (No Project Location specified.)
#       Add Preferred Skills column to filters (but not re drop).
#       Where possible drop named indexes (e.g. ix_my_actn) and replace with lookup of role_headers
#       Allow multiple values (in Mandatory Skills) to map to same filter value in My Filter.
#       Add option to delete rows with blank filter (for Mario).
#       Fix problem of deleted rows.  After the deletes, Excel still thinks that it has the original number of rows.
#       When reading dates from Action File, add test to make sure have actual date (re .strftime()).
#       Allow wild cards when specifying location.

import sys
import shlex
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils  import get_column_letter
import ReadControl
from FilterRow import FilterRow

arguments = sys.argv[0:]  # List of all arguments.  arguments[0] = <script name>  arguments[1] = <Control File name>   arguments[2] = <Input Spreadsheet File name>

now_time = datetime.now()
start_message = "Python script " + arguments[0] + " started at " + now_time.strftime("%Y-%m-%d %H:%M:%S")

# MessageOut(), MessageClose(), and MessageShow() functions for output message handling. [Development note: Should add code to handle exceptions.]
mssg_file = None
mssg_file_path = None
def MessageOut(mssg_txt) :  # Write message to output message file.
    global mssg_file, mssg_file_path
    if not mssg_file :
       mssg_file = open(mssg_file_path, "w")
    mssg_file.write(f"{mssg_txt}\n")
def MessageClose() :
    global mssg_file, mssg_file_path
    if mssg_file :
        MessageOut(" ")
        MessageShow("Output messages to: " + mssg_file_path)
        mssg_file.close()
def MessageShow(mssg_txt) :  # Display message to console and write to output message file.
    print(mssg_txt)
    MessageOut(mssg_txt)

if len(arguments) < 3 :
    print("Missing required command line parameters.")
    print(arguments)
    print(f"Usage: python {sys.argv[0]} <control file> <workbook>")
    print(" .. or, if using BAT file: MySource.bat <workbook>")
    print("Terminating process.")
    sys.exit(1)

# print("Python executable: ", sys.executable)
try:
    import pandas as pd
#   print("Pandas location:", pd.__file__)
#   print("Pandas version:", pd.__version__)
except ModuleNotFoundError:
    print("Pandas not found!")
    print("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has Pandas installed.")
    print("Terminating process.")
    sys.exit(1)
    
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("openpyxl not found!")
    print("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has openpyxl installed.")
    print("Terminating process.")
    sys.exit(1)

# Read Parameter File and load values into a dictionary object.
err_code, parms_dict = ReadControl.read(arguments[1], True)
if err_code > 0 :
    print("Terminating process.")
    sys.exit(1)

# Write start message to output message file.
mssg_file_path = parms_dict['messagdir'][0] + "\\MySource.mssg"  # Message Directory obtained from Parameter File.
MessageOut(start_message)
MessageOut(f"Using {arguments[1]} Control File.")
MessageOut(" ")

# Prepare array for processing row drops.
drop_dict = {}
drop_ix = 0
while drop_ix < len(parms_dict['droprows']) :
    drop_key = parms_dict['droprows'][drop_ix]
    drop_dict[drop_key] = [-1, 0, -1]  # Initialize the list with column index = -1, count of lines dropped = 0, xref = -1
    drop_ix +=1
 
do_actual_delete = False
early_quit = False
try:
    if parms_dict['dropactual'][0] == "True" :
        do_actual_delete = True
    elif parms_dict['dropactual'][0] == "Test" :
        early_quit = True
except KeyError:
    MessageOut(f"Did not locate \"dropactual\" parameter in {arguments[1]} Control File.  Default to \"False\".")
if do_actual_delete :
    MessageShow("WILL BE deleting rows during this run.")
else :
    MessageShow("WILL NOT BE deleting rows during this run.")

have_filtersheet = False
if "filtersheet" in parms_dict :
    if len(parms_dict['filtersheet']) < 2 :
        MessageShow(f"Missing required arguments from \"filtersheet\" parameter in {arguments[1]} Control File.")
        MessageShow("Should be: filtersheet <workbook path> <spreadsheet name>")
    else :
        filter_path = parms_dict['filtersheet'][0]
        filter_sheet = parms_dict['filtersheet'][1]
        MessageShow("filter_path = " + filter_path + ", filter_sheet = " + filter_sheet)
        have_filtersheet = True
else :
    MessageShow(f"\"filtersheet\" parameter missing from {arguments[1]} Control File.")

if not have_filtersheet :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

filter_object = FilterRow(filter_path, filter_sheet);
    
if early_quit:
    MessageShow("Test run - terminating here.")
    MessageClose()
    sys.exit(1)

# findActionHeaders() function - to identify column headers in Action File (worksheet)
ix_act_req = -1
ix_act_act = -1
ix_act_dte = -1
def findActionHeaders(header_row) :
    global ix_act_req, ix_act_act, ix_act_dte
    col_ix = 0
    while col_ix < len (header_row) :
        if   header_row[col_ix] == parms_dict['col_act_rqust'][0]:
            ix_act_req = col_ix
            MessageOut(f"{header_row[col_ix]} = {col_ix}")
        elif header_row[col_ix] == parms_dict['col_action'][0] :
            ix_act_act = col_ix
            MessageOut(f"{header_row[col_ix]} = {col_ix}")
        elif header_row[col_ix] == parms_dict['col_action_dt'][0] :
            ix_act_dte = col_ix
            MessageOut(f"{header_row[col_ix]} = {col_ix}")
        col_ix += 1
        
    if ix_act_req < 0 or ix_act_act < 0 or ix_act_dte < 0 :
        MessageShow(f"Did not find all column headers in Action Worksheet ({ix_act_req}, {ix_act_act}, {ix_act_dte}")
        MessageShow("Terminating process.")
        MessageClose()
        sys.exit(1)

# Read Action File and load into a dictionary object.
actionfile_path = parms_dict['inputdir'][0] + "\\" + parms_dict['actionf'][0]
MessageShow("Reading " + actionfile_path)
try:
    actbook  = load_workbook(actionfile_path)
    actsheet = actbook.active  # Get the active (only) worksheet.
except FileNotFoundError:
    MessageShow(f"File {actionfile_path} not found.")
except Exception as err:
    MessageShow(f"Could not open the file: {err}")

first_row = True
action_dict = {}
for row in actsheet.iter_rows(min_row=1, max_row=actsheet.max_row, values_only=True):
    if first_row :
        findActionHeaders(row)
        first_row = False
    elif row[ix_act_req] is not None :
        action_dict[row[ix_act_req]] = f"{row[ix_act_act]} {row[ix_act_dte].strftime("%m/%d/%Y")}"  # Development note: May need to test re if cell is a date object.

MessageShow(f"Read {len(action_dict)} actions from {actionfile_path}.")



#
#
# Read Roles Worksheet
workbook_path = parms_dict['inputdir'][0] + "\\" + arguments[2]
MessageShow("Reading " + workbook_path)
workbook_out = workbook_path.replace(".xlsx", "_out.xlsx")

try:
    workbook  = load_workbook(workbook_path)
    worksheet = workbook.active  # Get the active (only) worksheet.
    worksheet_name = workbook.sheetnames[0]
    MessageShow(f"Processing \"{worksheet_name}\" worksheet.")
except FileNotFoundError:
    MessageShow(f"File {workbook_path} not found.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    

# Test if column header matches any of the "droprows" columns.
def dropColumnCheck(header_text, col_ix) :
    global drop_dict, parms_dict
    drop_ix = 0
    drop_ix_max = len(parms_dict['droprows'])
    drop_ix_save = -1
    while drop_ix < drop_ix_max :
        drop_key = parms_dict['droprows'][drop_ix]
        if parms_dict[drop_key][0] == header_text :
            MessageOut (f"Index of \"{header_text}\" = {col_ix}  <- (drop test column)")
            drop_dict[drop_key] = [col_ix, 0, drop_ix]
        drop_ix +=1

# findRoleHeaders() function - to identify column headers in Role File (worksheet)
ix_request = -1
ix_my_actn = -1
ix_my_fltr = -1
ix_colocat = -1
ix_mylocat = -1
ix_mynsert = -1
role_headers = {}
def findRoleHeaders(header_row) :
    global ix_request, ix_my_actn, ix_my_fltr, ix_colocat, ix_mylocat, ix_mynsert, parms_dict, workbook_path, role_headers
 
    for idx, cell in enumerate(header_row):
        if cell.value is not None :
            col_header = cell.value.strip() 
            role_headers[col_header] = idx
    
    col_ix = 0
    while col_ix < len (header_row) :
        header_text = header_row[col_ix].value
        if header_text == parms_dict['col_request'][0] :    
            ix_request = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")
        elif header_text == parms_dict['col_my_act'][0] :
            ix_my_actn = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")    
        elif header_text == parms_dict['col_my_filter'][0] :
            ix_my_fltr = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")
        elif header_text == parms_dict['col_colocation'][0] :
            ix_colocat = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")
        elif header_text == parms_dict['col_mylocation'][0] :
            ix_mylocat = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")
        elif header_text == parms_dict['col_my_insert'][0] :
            ix_mynsert = col_ix
            MessageOut(f"Index of \"{header_text}\" = {col_ix}")
        
        dropColumnCheck(header_text, col_ix)        # Possible for some columns to be identified individually (above) and also in the "droprows" parameter. 
            
        col_ix += 1
        
    missing_header = False
    if ix_request < 0 :
        missing_header = True
        MessageShow(f"Did not find \"{parms_dict['col_request'][0]}\" column header in Role Worksheet, {workbook_path}.")
    if ix_mynsert < 0 :
        if ix_my_actn < 0 or ix_my_fltr < 0 :
            missing_header = True
            MessageShow(f"Did not find \"{parms_dict['col_my_insert'][0]}\" column header in Role Worksheet, {workbook_path}.  (To specify where to insert columns.)")
    if ix_my_actn < 0 :
        # Will insert column - so OK to continue processing.
        MessageOut(f"Did not find \"{parms_dict['col_my_act'][0]}\" column header in Role Worksheet, {workbook_path}.")
    if ix_my_fltr < 0 :
        # Will insert column - so OK to continue processing.
        MessageOut(f"Did not find \"{parms_dict['col_my_filter'][0]}\" column header in Role Worksheet, {workbook_path}.")
    if ix_colocat < 0 :
        missing_header = True
        MessageShow(f"Did not find \"{parms_dict['col_colocation'][0]}\" column header in Role Worksheet, {workbook_path}.")
    if ix_mylocat < 0 :
        missing_header = True
        MessageShow(f"Did not find \"{parms_dict['col_mylocation'][0]}\" column header in Role Worksheet, {workbook_path}.")
    if missing_header :
        MessageShow("Terminating process.")
        MessageClose()
        sys.exit(1)


#   matchDrop()- Compare the current cell value against any of the specified "drop" or "keep", or other tests.
#   This function tests only a single cell of the current row, against a single condition from "droprows".  The condition definition specifies which cell in the row is being tested.
#
#   Control File Format (First element returned as parms_dict key, subsequent elements returned in list object in parms_dict.):
#
#           droprows condition1 condition2 condition3 ...
#           condition1 "column header 1" drop dropvalue1 dropvalue2 dropvalue3 ...
#           condition2 "column header 2" keep keepvalue1 keepvalue2 keepvalue3 ...
#
#   Condition types:
#   
#       drop    - Drop the row if the cell value matches any of the values listed for the condition.
#       dropbl  - Same as "drop" but also if the cell value is blank or empty.
#       dropincl - Drop the row if the cell value includes any of the values listed for the condition.  (Different from "drop" b/c "drop" test if equal to the whole cell value.)
#       keep    - Keep the row if the cell value matches any of the values listed for the condition.
#       keepbl  - Same as "keep" but also if the cell value is blank or empty.
#       before  - Drop the row if the date is not before the specified date.
#       location - Compares two cells of the row (My Location and Co-Location).  Here, we just set the test_colocation flag.  Actual comparison takes place in the colocTooFar() function.
#       filter  - Search for a string within the cell value.  If found, set the my_filter variable and found_filter flag.
#       nofilter - Drop the row if there is no filter set.  This condition should be listed after any "filter" conditions in the droprows arguments.


def matchDrop(cell_value, xref_ix) :
    global parms_dict, my_filter, found_filter, test_colocation
    
    drop_key = parms_dict['droprows'][xref_ix]
    is_date_instance = False
    test_colocation  = False

    if isinstance(cell_value, datetime):    # do this to avoid hitting error on cell_value.strip() - in the else construct.
        is_date_instance = True

    elif cell_value is None or cell_value.strip() == "" :
        if  parms_dict[drop_key][1] == "dropbl" :
            return True
        elif parms_dict[drop_key][1] == "keepbl" :
            return False
        elif parms_dict[drop_key][1] == "filter" :
            return False
        elif parms_dict[drop_key][1] == "dropincl" :
            return False
    
    col_ix = 2
    if  parms_dict[drop_key][1] == "drop" or parms_dict[drop_key][1] == "dropbl":
        while col_ix < len (parms_dict[drop_key]) :
            if parms_dict[drop_key][col_ix] == cell_value :
                return True
            col_ix += 1
        return False
        
    elif parms_dict[drop_key][1] == "keep" or parms_dict[drop_key][1] == "keepbl":
        while col_ix < len (parms_dict[drop_key]) :
            if parms_dict[drop_key][col_ix] == cell_value :   
                return False
            col_ix += 1
        return True
        
    elif parms_dict[drop_key][1] == "before" :
        if is_date_instance :
            test_date = datetime.strptime(parms_dict[drop_key][2], "%Y%m%d")
            if cell_value < test_date :
                return False
            else :
                return True
        else :
            return False
            
    elif parms_dict[drop_key][1] == "location" :    # Requires comparison of two cells - so set the test_colocation flag & pass control back to dropThisRow()
        test_colocation = True
        return False
            
    elif parms_dict[drop_key][1] == "filter" :   # "filter" action is to find any of the filter values in the cell_value.
        if found_filter :
            return False    # Bypass this test if My Filter has already been set for this row.
        while col_ix < len (parms_dict[drop_key]) :
            if parms_dict[drop_key][col_ix].lower() in cell_value.lower() :
                my_filter = parms_dict[drop_key][col_ix]
                found_filter = True
                return False
            col_ix += 1
        return False

    elif parms_dict[drop_key][1] == "nofilter" :    # Special condition - drop this row if filter is blank.
    #   This drop_key should follow any of the filter drop_keys in the "droprows" list. 
        if found_filter :
            return False
        else :
            return True
        
    elif parms_dict[drop_key][1] == "dropincl" : # Drop if any of the drop values are included in the cell_value.
        if found_filter :
            return False    # Bypass this test if My Filter has already been set for this row.
        while col_ix < len (parms_dict[drop_key]) :
            if parms_dict[drop_key][col_ix].lower() in cell_value.lower() :
               return True
            col_ix += 1
        return False
    
    return False    # Just in case - processing should not reach this line of code.


my_non_loc_count = 0
#   Compare Co-Location requirement against "My Location".
def colocTooFar(xref_ix, co_location, my_location) :
    global parms_dict, my_non_loc_count
    drop_key  = parms_dict['droprows'][xref_ix]
    coloc_key = parms_dict[drop_key][2]
    myloc_key = parms_dict[drop_key][3]
    
    if my_location is None or my_location == "" :
        MessageShow("Blank My Location value.  Should have been filled by MyLocation.py.")
        my_non_loc_count += 1
        if my_non_loc_count > 10 :
            MessageShow("Too many blank My Location values.")
            MessageShow("Terminating process.")
            MessageClose()
            sys.exit(1)
        return False
    
    coloc_pct = -1
    for loc_key in parms_dict[coloc_key] :
        if parms_dict[loc_key][1] == co_location :
            coloc_pct = int(parms_dict[loc_key][0])
    if coloc_pct < 0 :
        MessageShow(f"Did not find match for \"{co_location}\" in {coloc_key}.")
        
    myloc_pct = -1
    for loc_key in parms_dict[myloc_key] :
        loc_ix = 1  # Start with 1 b/c index 0 is percentage.
        while loc_ix < len (parms_dict[loc_key]) :
            if parms_dict[loc_key][loc_ix] == my_location :
                myloc_pct = int(parms_dict[loc_key][0])
            loc_ix += 1
    if myloc_pct < 0 :
        MessageShow(f"Did not find match for \"{my_location}\" in {myloc_key}.  Note: Is case-sensitive.")
        
    if myloc_pct < coloc_pct:
        return True

    return False


#   Test if current row should be dropped.
def dropThisRow(test_row) :
    global drop_dict, test_colocation, ix_colocat, ix_mylocat


    for drop_line in drop_dict :   # drop_dict contains the column number of all of the columns that are tested.
        test_col_ix = drop_dict[drop_line][0]
        
        if test_col_ix > -1 :    
            test_colocation = False # Flag Co-Location test, since cannot be processed in MatchDrop().  Will call colocTooFar() instead - see 5 lines down.

            if matchDrop (test_row[test_col_ix].value, drop_dict[drop_line][2]) :   # Function call arguments: matchDrop(<cell value>, <xref_ix>)
                drop_dict[drop_line][1] += 1
                return True
            
            if test_colocation :
                if colocTooFar (drop_dict[drop_line][2], test_row[ix_colocat].value, test_row[ix_mylocat].value) :
                    drop_dict[drop_line][1] += 1
                    return True

    return False
    

    
#   Process Roles Worksheet

#   If skip_to_header parameter specified, look for the specified column title to identify the header row.  Otherwise, assume the first row of the worksheet is the header.
have_header_row    = False
start_header_value = ""
start_header_index = 0
if "skip_to_header" in parms_dict :
    if len(parms_dict['skip_to_header']) > 0 :
        start_header_value = parms_dict['skip_to_header'][0]
    if len(parms_dict['skip_to_header']) > 1 :
        index_number = parms_dict['skip_to_header'][1]
        try :
            start_header_index = int(index_number)
        except ValueError :
            start_header_index = -1
        if start_header_index < 0 :
            MessageShow(f"Expected a positive number in the second argument of \"skip_to_header\" parameter, but found\"{index_number}\"  Defaulting to zero.")
            start_header_index = 0
        
count_match = 0
count_nomat = 0
count_nofilter =0
count_delet = 0
process_count = 0
rows_to_delete = []
my_filter = ""
min_data_row = 1

for row_ix, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=False), start=1) :
    if not have_header_row :
        if start_header_value == "" :
            have_header_row = True
        elif start_header_value == row[start_header_index].value :
            have_header_row = True
        if have_header_row :
            findRoleHeaders(row)
            min_data_row = row_ix + 1

if not have_header_row :
    MessageShow(f"Unable to locate the header row in \"{workbook.sheetnames[0]}\" worksheet.  Searched first 12 rows for \"{start_header_value}\" in column {start_header_index}.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
           
def formatColumnText(col_title) :
    global role_headers, worksheet_name, min_data_row, worksheet
    if col_title not in role_headers :
        MessageShow(f"Did not find column title, \"{col_title}\" in the \"{worksheet_name}\" worksheet header.")
        MessageShow(f"The \"{col_title}\" title was specified in the \"format_cols\" parameter.")
        return
        
    left_wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    col_letter = get_column_letter(role_headers[col_title] + 1)     # Add 1 to the column index to point correctly.
    MessageShow(f"Formatting text alignment in column \"{col_title}\" ({col_letter}).")
    for row in range(min_data_row, worksheet.max_row + 1):
        cell = worksheet[f'{col_letter}{row}']
        cell.alignment = left_wrap_alignment
        worksheet.row_dimensions[row].height = 15

#
#   If inserting a column, make sure all column indexes greater than the insert columns are incremented.
def bumpColIndexes(bump_ix) :
    global ix_request, ix_colocat, ix_mylocat, drop_dict, role_headers
    
    if bump_ix < ix_request :
        ix_request += 1
    if bump_ix < ix_colocat :
        ix_colocat += 1
    if bump_ix < ix_mylocat :
        ix_mylocat += 1
        
    for drop_line in drop_dict :   # drop_dict contains the column number of all of the columns that are tested.
        if bump_ix <= drop_dict[drop_line][0] :
            drop_dict[drop_line][0] += 1
    
    for key, value in role_headers.items():
        if bump_ix <= value :
            role_headers[key] += 1
            
if ix_my_actn < 0 :
    column_title = parms_dict['col_my_act'][0]
    worksheet.insert_cols(ix_mynsert + 2)   # Count plus two from specified column - to insert to right of specified column.
    ix_my_actn = ix_mynsert + 1             # Count plus one from specified column - for index of new column.
    worksheet.cell(row=min_data_row - 1, column=ix_my_actn + 1, value=column_title)    # Write column header
    MessageShow(f"Inserting \"{parms_dict['col_my_act'][0]}\" column in \"{workbook.sheetnames[0]}\" worksheet.")
    MessageOut(f"ix_my_actn = {ix_my_actn}")
    bumpColIndexes(ix_my_actn)
    role_headers[column_title] = ix_my_actn
if ix_my_fltr < 0 :
    column_title = parms_dict['col_my_filter'][0]
    worksheet.insert_cols(ix_my_actn + 2)   # Count plus two from specified column - to insert to right of specified column.
    ix_my_fltr = ix_my_actn + 1             # Count plus one from specified column - for index of new column.
    worksheet.cell(row=min_data_row - 1, column=ix_my_fltr + 1, value=column_title)    # Write column header
    MessageShow(f"Inserting \"{parms_dict['col_my_filter'][0]}\" column in \"{workbook.sheetnames[0]}\" worksheet.")
    MessageOut(f"ix_my_fltr = {ix_my_fltr}")
    bumpColIndexes(ix_my_fltr)
    role_headers[column_title] = ix_my_fltr


if "format_cols" in parms_dict :
    for col_title in parms_dict['format_cols'] :
        formatColumnText(col_title)


for row_ix, row in enumerate(worksheet.iter_rows(min_row=min_data_row, max_row=worksheet.max_row, values_only=False), start=1) :
         
    if row[ix_request].value is not None :
        this_request_value = row[ix_request].value
        if isinstance(this_request_value, str) :
            this_request_id = int(this_request_value)
        else :
            this_request_id = this_request_value
            
        my_filter = ""
        found_filter = False
        increment_blank = 0
        if dropThisRow(row) :
            count_delet += 1
            if do_actual_delete :
                rows_to_delete.append(row_ix + min_data_row - 1)
            else :
                my_filter = "DELETE"
                found_filter = True
        elif this_request_id in action_dict :
            count_match += 1
            row[ix_my_actn].value = action_dict[this_request_id]
            increment_blank = 1
        else :
            count_nomat += 1
            increment_blank = 1
        
        if found_filter :
            row[ix_my_fltr].value = my_filter
        else :
            count_nofilter += increment_blank

        process_count +=1
        if process_count > 199 :
            print("processing ....")
            process_count = 0

if do_actual_delete :
    MessageShow(f"Dropping {count_delet} rows.")            
    deleting_count = 0
    delete_to_go = count_delet
    for row in reversed(rows_to_delete) :
        worksheet.delete_rows(row, 1)
    
        deleting_count +=1
        delete_to_go -= 1
        if deleting_count > 99 :
            print(f"deleting ....... {delete_to_go}")
            deleting_count = 0
else :
    MessageShow(f"No rows actually deleted during this run. {count_delet} rows marked with \"DELETE\" in \"{parms_dict['col_my_filter'][0]}\" column.")
        
workbook.save(workbook_out)

for drop_line in drop_dict :
    drop_tally = drop_dict[drop_line][1]
    drop_key_ix = drop_dict[drop_line][2]
    drop_parm_ix = parms_dict['droprows'][drop_key_ix]
    drop_header = parms_dict[drop_parm_ix][0]
    if drop_tally > 0 :
        MessageShow(f"Dropped {drop_tally} rows b/c \"{drop_header}\" value.")
        
MessageShow(f"{count_match} rows matched, {count_nomat} rows not matched, from {count_match + count_nomat} output rows")
MessageShow(f"{count_nofilter} output rows with blank in \"{parms_dict['col_my_filter'][0]}\" column.")
MessageShow(f"Worksheet saved to {workbook_out}.")
MessageClose()