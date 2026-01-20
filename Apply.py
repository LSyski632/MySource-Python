# Apply.py - Provide easy way to update action spreadsheet for roles applied for or passed over.

# Suggested Enhancements:
#       Make sure target workbook is closed before doing final save.  (How this done?)

import sys
import shlex
from pathlib import Path
from datetime import datetime
import ReadControl

arguments = sys.argv[0:]  # List of all arguments.  
# arguments[0] = <script name>  arguments[1] = <Control File name>  arguments[2] = <Input worksheet File name>
now_time = datetime.now()
start_message = "Python script " + arguments[0] + " started at " + now_time.strftime("%Y-%m-%d %H:%M:%S")

if len(arguments) < 3 :
    print("Missing required command line parameters.")
    print(arguments)
    print(f"Usage: python {sys.argv[0]} <control file> <workbook>")
    print(" .. or, if using BAT file: Apply.bat <workbook>")
    print("Terminating process.")
    sys.exit(1)
    
worksheet_path = Path(arguments[2])
if not worksheet_path.is_file():
    print(f"Input worksheet \"{arguments[2]}\" not found.")
    print("Terminating process.")
    sys.exit(1)
    
# Read parameter file and load values into a dictionary object.
err_code, parms_dict = ReadControl.read(arguments[1], True)
if err_code > 0 :
    print("Terminating process.")
    sys.exit(1)

# MessageOut(), MessageClose(), and MessageShow() functions for output message handling. [Development note: Should add code to handle exceptions.]
mssg_file = None
mssg_file_path = None
def MessageOut(mssg_txt) :  # Write message to output message file.
    global mssg_file, mssg_file_path
    if not mssg_file :
       mssg_file = open(mssg_file_path, "w")
    mssg_file.write(f"{mssg_txt}\n")
def MessageClose() :
    global mssg_file, mssg_file_path
    if mssg_file :
        MessageOut(" ")
        MessageShow("Output messages to: " + mssg_file_path)
        mssg_file.close()
def MessageShow(mssg_txt) :  # Display message to console and write to output message file.
    print(mssg_txt)
    MessageOut(mssg_txt)
    
# Write start message to output message file.
if 'messagdir' not in parms_dict:
    print(f"\"messagdir\" parameter not found in {arguments[1]} Control File.")
    print("Terminating process.")
    sys.exit(1)
    
mssg_file_path = parms_dict['messagdir'][0] + "\\Apply.mssg"  # Message Directory obtained from Parameter File.
MessageOut(start_message)
MessageOut(" ")

have_other_parms = True
#
# Test for required pandas and openpyxl libraries in current environment.
try:
    import pandas as pd
#   print("Pandas location:", pd.__file__)
#   print("Pandas version:", pd.__version__)
except ModuleNotFoundError:
    MessageShow("Pandas not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has Pandas installed.")
    have_other_parms = False
    
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    MessageShow("openpyxl not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has openpyxl installed.")
    have_other_parms = False

if 'inputdir' not in parms_dict:
    MessageShow(f"\"inputdir\" parameter not found in {arguments[1]} Control File.")
    have_other_parms = False
elif len(parms_dict['inputdir']) < 1 :
    MessageShow(f"\"inputdir\" parameter in {arguments[1]} Control File should have an argument: <workbook file name>")
    have_other_parms = False

if 'targetsheet' not in parms_dict:
    MessageShow(f"\"targetsheet\" parameter not found in {arguments[1]} Control File.")
    have_other_parms = False
elif len(parms_dict['targetsheet']) < 1 :
    MessageShow(f"\"targetsheet\" parameter in {arguments[1]} Control File should have an argument: <workbook file name>")
    have_other_parms = False
    
if 'copycols' not in parms_dict:
    MessageShow(f"\"copycols\" parameter not found in {arguments[1]} Control File.")
    have_other_parms = False
elif len(parms_dict['copycols']) < 2 :
    MessageShow(f"\"copycols\" parameter in {arguments[1]} Control File should have two arguments: <workbook file name> <worksheet name>")
    have_other_parms = False
    
if not have_other_parms :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
#
#
# Read Control Worksheet
controlbook_path  = parms_dict['copycols'][0]
controlsheet_name = parms_dict['copycols'][1]
have_controlsheet = True

try:
    controlbook  = load_workbook(controlbook_path)
    if controlsheet_name in controlbook.sheetnames:
        controlsheet = controlbook[controlsheet_name]
        MessageShow(f"Reading \"{controlsheet_name}\" worksheet in {controlbook_path}")
    else :
        MessageShow(f"Did not locate \"{controlsheet_name}\" worksheet in {controlbook_path}")
        have_controlsheet = False

except FileNotFoundError:
    MessageShow(f"File {controlbook_path} not found.")
    have_controlsheet = False
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    have_controlsheet = False
 
if not have_controlsheet :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
#   Find column headers in control worksheet.
ctrl_headers = {}
for idx, cell in enumerate(controlsheet[1]):
    if cell.value is not None :
        col_header = cell.value.strip() 
        ctrl_headers[col_header] = idx

missing_required_header = False
required_headers = ['ID', 'Source', 'Target', 'Type', 'Display', 'Apply', 'Pass', 'Query', 'Prompt']
for r_header in required_headers :
    if r_header not in ctrl_headers :
        MessageShow(f"Did not find requred column header \"{r_header}\" on {controlsheet_name} worksheet.")
        missing_required_header = True
if missing_required_header :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
#   Buld dictionary of column codes.
id_cix = ctrl_headers['ID']
ctrl_codes = {}
ctrl_ix = 1
for c_row in controlsheet.iter_rows(min_row=2, max_row=controlsheet.max_row, values_only=False) :
    ctrl_ix += 1
    if c_row[id_cix].value is not None :
        ctrl_codes[c_row[id_cix].value] = ctrl_ix
    
#
#
# Read Roles Worksheet
workbook_path = parms_dict['inputdir'][0] + "\\" + arguments[2]
MessageShow("Reading " + workbook_path)

try:
    workbook  = load_workbook(workbook_path)
    worksheet = workbook.active  # Get the active (only) worksheet.
    worksheet_name = workbook.sheetnames[0] 
except FileNotFoundError:
    MessageShow(f"File {workbook_path} not found.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

#   Find column headers.
role_headers = {}
def getRoleHeaders (sheet_row) :
    global role_headers
    for idx, cell in enumerate(sheet_row):
        if cell.value is not None :
            col_header = cell.value.strip() 
            role_headers[col_header] = idx
            
start_data_row = 2
getRoleHeaders(worksheet[1])
if "skip_header" in parms_dict :    # Special processing for GPS_Open_Demands_Report.xlsx - headers needed for processing are on second line.
    skip_header = parms_dict['skip_header'][0]
    if skip_header in role_headers and role_headers[skip_header] == 0 :
        role_headers = {}
        start_data_row += 1 
        getRoleHeaders(worksheet[2])

def getIndexOfColumn (column_title) :
    global worksheet_name, role_headers
    if column_title in role_headers :
        ix_request = role_headers[column_title]
    else:
        MessageShow(f"Did not locate \"{column_title}\" column header in \"{worksheet_name}\" worksheet.")
        ix_request = -1
    return ix_request

#   getColHeader() - Find a parameter in the Control File.  The code passed to this function is a parameter that should point to another parameter in the Control file, with two arguments. 
def getColHeader(col_parm_code) :
    global arguments, parms_dict, ctrl_codes, ctrl_headers, controlsheet
    
    if col_parm_code not in parms_dict :
        MessageShow(f"\"{col_parm_code}\" parameter is not specified in the {arguments[1]} Control File.")
        return -1
    
    if len(parms_dict[col_parm_code]) < 1 :
        MessageShow(f"No value is specified for the \"{col_parm_code}\" parameter in the {arguments[1]} Control File.")
        return -1
        
    col_reference = parms_dict[col_parm_code][0]
    if col_reference not in ctrl_codes :
        MessageShow(f"\"{col_reference}\" parameter is not specified in the {controlsheet_name} Control Worksheet.")
        MessageShow(f"The \"{col_reference}\" pointer was specified in the \"{col_parm_code}\" parameter.")
        return -1
        
    source_cix = ctrl_headers['Source']
    c_row_index = ctrl_codes[col_reference]
    return_index = getIndexOfColumn(controlsheet[c_row_index][source_cix].value)
    return return_index
    
ix_request = getColHeader("col_request")
ix_role    = getColHeader("col_role")
if ix_request < 0 or ix_role < 0 :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
# Save row indexes into a Python dictionary.
request_dict = {}
row_count = start_data_row - 1
blank_rows = 0
for row in worksheet.iter_rows(min_row=start_data_row, max_row=worksheet.max_row, values_only=False):
    row_count +=1
    if row[ix_request].value is None :  # Test for blank rows, and ignore anything beyond the first couple of blank Request IDs.
                                # Excel seems to be confused by the MySource process, and does not know that rows have been deleted.  After going through the real rows, worksheet.iter_rows(...) 
                                # keeps going and makes repeated passes through the actual rows. This would throw off the row_count value, so that later retrieval would not work.
        blank_rows += 1
    elif blank_rows < 3 :
        request_id_value = row[ix_request].value
        if isinstance(request_id_value, int) :
            request_dict[request_id_value] = row_count
        elif isinstance(request_id_value, str) :
            if len(request_id_value) > 6 :
                print (request_id_value)
                request_id_value = request_id_value[-6:]
            try :
                request_id_int = int(request_id_value)
                request_dict[request_id_int] = row_count
            except ValueError:
                MessageShow(f"Request ID not valid; found \"{request_id_value}\"")
                
MessageShow(f"Read {len(request_dict)} rows from \"{worksheet_name}\" worksheet.")

#
#
# Read Target Worksheet
target_path = parms_dict['targetsheet'][0]
MessageShow("Reading " + target_path)

try:
    targetbook  = load_workbook(target_path)
    targetsheet = targetbook.active  # Get the active (only) worksheet.
    targetsheet_name = targetbook.sheetnames[0] 
except FileNotFoundError:
    MessageShow(f"File {target_path} not found.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

#   Find target column headers.
target_headers = {}
max_target_col = 0
for idx, cell in enumerate(targetsheet[1]):
    if cell.value is not None :
        col_header = cell.value.strip() 
        target_headers[col_header] = idx
        max_target_col = idx

#   Test input Request ID format and value.  This value is entered by the user, from a prompt.
def testRequestInput(string_input) :
    try:
        num_value = int(string_input)
    except ValueError:
        MessageShow(f"Invalid Request ID.  Should be a six-digit number, found \"{string_input}\".")
        return -1
    if num_value > 999999 :
        MessageShow(f"Invalid Request ID.  Should be a six-digit number, found {num_value}.")
        return -1
    if num_value < 1 :
        MessageShow(f"Invalid Request ID.  Should be strictly positive, found {num_value}.")
    return num_value
    
def displayRowInfo(search_request_id) :
    global request_dict, worksheet, ix_request, ix_role, target_row, current_request_tag, controlsheet
    target_row_num = request_dict[search_request_id]
    target_row = worksheet[target_row_num]
    
    current_request_tag = f"Request # {target_row[ix_request].value}, {target_row[ix_role].value}"
    MessageShow("")
    MessageShow(current_request_tag)
    
    display_cix = ctrl_headers['Display']
    source_cix  = ctrl_headers['Source']
    target_cix  = ctrl_headers['Target']
    for c_row in controlsheet.iter_rows(min_row=2, max_row=controlsheet.max_row, values_only=False) :
        if c_row[display_cix].value is not None :
            source_col_name = c_row[source_cix].value
            # print(f"source_col_name = {source_col_name}")
            ix_source = getIndexOfColumn(source_col_name)
            if ix_source > -1 :
                MessageShow(f"{c_row[target_cix].value}: {target_row[ix_source].value}")

#         
#   inputAction() - Get user response re Apply For or Pass Over - re current Request ID.  Update target worksheet with action information for the current Request ID.
def inputAction(request_id) :
    global target_path, targetsheet, current_request_tag, target_updated, col_types
    print (" "  )
    input_action = input("Pass Over or Apply For this role? ")
    if input_action == "" :
        MessageShow(f"No action taken for Request {request_id}.")
        return
    upper_action = input_action.capitalize()
    if upper_action[0:1] == "A" :
        action_name = "Applied"
        new_row = buildRow("Applied", request_id)
    elif upper_action[0:1] == "P" :
        action_name = "Pass"
        new_row = buildRow("Pass", request_id)
    else:
        MessageShow(f"No action taken for Request {request_id}.")
        return
        
    if isinstance(new_row, list) :
        targetsheet.append(new_row)
        target_updated = True
        for col_ix, col_value in enumerate(col_types):
            if col_value == "date" :
                date_cell = targetsheet.cell(row=targetsheet.max_row, column=(col_ix + 1))
                date_cell.number_format = 'mm/dd/yyyy'
                
        MessageShow(f"Updated {target_path} with (\"{action_name}\") for {current_request_tag}.")
    
#
#   buildRow() - Prepare list object containing elements to be appended to the target worksheet row.
def buildRow(action, request_id) :
    global arguments, parms_dict, max_target_col, target_row, col_types, ctrl_headers, controlsheet
    new_row = [""] * (max_target_col + 1)
    col_types = ["str"] * (max_target_col + 1)
    if action == "Applied" :
        action_cix = ctrl_headers['Apply']
    elif action == "Pass" :
        action_cix = ctrl_headers['Pass']
    else :
        return None
        
    source_cix  = ctrl_headers['Source']
    target_cix  = ctrl_headers['Target']
    type_cix    = ctrl_headers['Type']
    query_cix   = ctrl_headers['Query']
    prompt_cix  = ctrl_headers['Prompt']
    for c_row in controlsheet.iter_rows(min_row=2, max_row=controlsheet.max_row, values_only=False) :
        if c_row[action_cix].value is not None :
            source_col_name = c_row[source_cix].value
            target_col_name = c_row[target_cix].value
            col_index = getTargetIndex(target_col_name)
            if col_index > -1 :
                if source_col_name == "?" :
                    query_type = c_row[query_cix].value
                    if query_type == "today" :
                        todays_date = datetime.now()
                        new_row[col_index] = todays_date
                        col_types[col_index] = "date"
                    elif query_type == "action" :
                        new_row[col_index] = action
                    elif query_type == "prompt" :
                        input_info = input(c_row[prompt_cix].value + " ")
                        new_row[col_index] = input_info
                    elif query_type == "literal" :
                        new_row[col_index] = c_row[prompt_cix].value
                    else :
                        MessageShow(f"*** Note: \"{query_type}\" argument for \"{target_col_name}\" parameter not understood.")
                else : 
                    ix_source = getIndexOfColumn(source_col_name)
                    if ix_source > -1 :
                        col_types[col_index] = c_row[type_cix].value
                        if col_types[col_index] == "int" :
                            try :
                                new_row[col_index] = int(target_row[ix_source].value)
                            except ValueError :
                                MessageShow(f"Unexpected value, \"{target_row[ix_source].value}\" for {target_col_name}.")
                        else :
                            new_row[col_index] = target_row[ix_source].value                            
    return new_row
    

def getTargetIndex (column_title) :
    global target_headers, targetsheet_name
    if column_title in target_headers :
        ix_request = target_headers[column_title]
    else:
        MessageShow(f"Did not locate \"{column_title}\" column header in \"{targetsheet_name}\" worksheet.")
        ix_request = -1
    return ix_request
    
target_updated = False
while True :
    print ("")
    input_request_id = input("Please enter Request ID: ")
    if input_request_id == "" :
        print("")
        MessageShow("Ending process.")
        break
    search_request_id = testRequestInput(input_request_id)
    if search_request_id > 0 :
        if search_request_id in request_dict :
            displayRowInfo(search_request_id)
            inputAction(search_request_id)
                
        else :
            MessageShow (f"Request ID {search_request_id} not found in \"{worksheet_name}\" worksheet.")

if target_updated :
    input_dummy = input(f"Is the {targetsheet_name} workbook closed?  If not, will CRASH !!!  <Enter> to continue.")
    targetbook.save(target_path)
    
MessageClose()
