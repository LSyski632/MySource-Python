# Mario.py - Populate Action File from prior filtered Open Roles worksheet.  The filtered worksheet includes a "My Action" column that specifies an action taken & date.

# This is used if the subject (Mario) does not mainitain his own Action File. 

# Suggested Enhancements:
#   Add routine to verify all required parameters are specified in the control file. (Pass list to ReadControl, with number of required parameters for each.)
#   Add processing for skip_to_header parmeter (as in MySource.py).
#   Provide graceful exit if worksheet is open in Excel session.

import sys
import shlex
from pathlib import Path
from datetime import datetime
import ReadControl

arguments = sys.argv[0:]  # List of all arguments.  arguments[0] = <script name>  arguments[1] = <Control File name>   arguments[2] = <Prior Open Roles worksheet file name>
now_time = datetime.now()
start_message = "Python script " + arguments[0] + " started at " + now_time.strftime("%Y-%m-%d %H:%M:%S")

if len(arguments) < 3 :
    print("Missing required command line parameters.")
    print(arguments)
    print(f"Usage: python {sys.argv[0]} <control file> <workbook>")
    print(" .. or, if using BAT file: Mario.bat <workbook>")
    print("Terminating process.")
    sys.exit(1)

worksheet_path = Path(arguments[2])
if not worksheet_path.is_file():
    print(f"Input worksheet \"{arguments[2]}\" not found.")
    print("Terminating process.")
    sys.exit(1)

# Read parameter file and load values into a dictionary object.
err_code, parms_dict = ReadControl.read(arguments[1], True)
if err_code > 0 :
    print("Terminating process.")
    sys.exit(1)

# MessageOut(), MessageClose(), and MessageShow() functions for output message handling. [Development note: Should add code to handle exceptions.]
mssg_file = None
mssg_file_path = None
def MessageOut(mssg_txt) :  # Write message to output message file.
    global mssg_file, mssg_file_path
    if not mssg_file :
       mssg_file = open(mssg_file_path, "w")
    mssg_file.write(f"{mssg_txt}\n")
def MessageClose() :
    global mssg_file, mssg_file_path
    if mssg_file :
        MessageOut(" ")
        MessageShow("Output messages to: " + mssg_file_path)
        mssg_file.close()
def MessageShow(mssg_txt) :  # Display message to console and write to output message file.
    print(mssg_txt)
    MessageOut(mssg_txt)
    
# Write start message to output message file.
mssg_file_path = parms_dict['messagdir'][0] + "\\Mario.mssg"  # Message Directory obtained from Parameter File.
MessageOut(start_message)
MessageOut(" ")

# print("Python executable: ", sys.executable)
try:
    import pandas as pd
#   print("Pandas location:", pd.__file__)
#   print("Pandas version:", pd.__version__)
except ModuleNotFoundError:
    MessageShow("Pandas not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has Pandas installed.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    MessageShow("openpyxl not found!")
    MessageShow("Make sure that you run " + arguments[0] + " in a virtual environment that is activated and has openpyxl installed.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    

#
#
# Read Control Worksheet
controlbook_path  = parms_dict['contrlsheet'][0]
controlsheet_name = parms_dict['contrlsheet'][1]
have_controlsheet = True

try:
    controlbook  = load_workbook(controlbook_path)
    if controlsheet_name in controlbook.sheetnames:
        controlsheet = controlbook[controlsheet_name]
        MessageShow(f"Reading \"{controlsheet_name}\" worksheet in {controlbook_path}")
    else :
        MessageShow(f"Did not locate \"{controlsheet_name}\" worksheet in {controlbook_path}")
        have_controlsheet = False

except FileNotFoundError:
    MessageShow(f"File {controlbook_path} not found.")
    have_controlsheet = False
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    have_controlsheet = False
 
if not have_controlsheet :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
#   Find column headers in Control Worksheet.  (Assume headers are in first row.)  Save column indexes into ctrl_headers dictionary.
ctrl_headers = {}
for idx, cell in enumerate(controlsheet[1]):
    if cell.value is not None :
        col_header = cell.value.strip() 
        ctrl_headers[col_header] = idx

missing_required_header = False
required_headers = ['Source', 'Target', 'Type', 'Process']
for r_header in required_headers :
    if r_header not in ctrl_headers :
        MessageShow(f"Did not find required column header \"{r_header}\" on {controlsheet_name} worksheet.")
        missing_required_header = True
if missing_required_header :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
# Read Control Worksheet and load into a dictionary of tuples. 
ctrl_dict = {}
for ctrl_row in controlsheet.iter_rows(min_row=2, max_row=controlsheet.max_row, values_only=False):
    ix_src  = ctrl_headers['Source']
    ix_targ = ctrl_headers['Target']
    ix_type = ctrl_headers['Type']
    ix_proc = ctrl_headers['Process']
    ctrl_dict[ctrl_row[ix_targ]] = (ctrl_row[ix_src], ctrl_row[ix_type], ctrl_row[ix_proc])

#
#
# Read Action File and process.
actionfile_path = parms_dict['actionsheet'][0] 
MessageShow("Reading " + actionfile_path)
try:
    actbook  = load_workbook(actionfile_path)
    actsheet = actbook.active  # Get the active (only) worksheet.
    actsheet_name = actbook.sheetnames[0]
except FileNotFoundError:
    MessageShow(f"File {actionfile_path} not found.")
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    
#   Find column headers in Action File (worksheet).  Headers should be in the first row.  Save column indexes into action_headers dictionary.
action_headers = {}
max_action_col_ix = -1
for idx, cell in enumerate(actsheet[1]):
    if cell.value is not None :
        col_header = cell.value.strip() 
        action_headers[col_header] = idx
        max_action_col_ix = idx
 
#
#
#   Process Source Worksheet
workbook_path = parms_dict['inputdir'][0] + "\\" + arguments[2]
MessageShow("Reading " + workbook_path)

try:
    workbook  = load_workbook(workbook_path)
    worksheet = workbook.active  # Get the active (only) worksheet.
    worksheet_name = workbook.sheetnames[0]
    MessageShow(f"Processing \"{worksheet_name}\" worksheet.")
except FileNotFoundError:
    MessageShow(f"File {workbook_path} not found.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
except Exception as err:
    MessageShow(f"Could not open the file: {err}")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)

#   If skip_to_header parameter specified, look for the specified column title to identify the header row.  Otherwise, assume the first row of the worksheet is the header.
have_header_row    = False
start_header_value = ""
start_header_index = 0
if "skip_to_header" in parms_dict :
    if len(parms_dict['skip_to_header']) > 0 :
        start_header_value = parms_dict['skip_to_header'][0]
    if len(parms_dict['skip_to_header']) > 1 :
        index_number = parms_dict['skip_to_header'][1]
        try :
            start_header_index = int(index_number)
        except ValueError :
            start_header_index = -1
        if start_header_index < 0 :
            MessageShow(f"Expected a positive number in the second argument of \"skip_to_header\" parameter, but found\"{index_number}\"  Defaulting to zero.")
            start_header_index = 0

# findRoleHeaders() function - to identify column headers in the Source Worksheet.
role_headers = {}
def findRoleHeaders(header_row) :
    global parms_dict, workbook_path, role_headers
    for idx, cell in enumerate(header_row):
        if cell.value is not None :
            col_header = cell.value.strip() 
            role_headers[col_header] = idx
            
min_data_row = 1
for row_ix, row in enumerate(worksheet.iter_rows(min_row=1, max_row=12, values_only=False), start=1) :
    if not have_header_row :
        if start_header_value == "" :
            have_header_row = True  # "skip_to_header" parameter not specified in Control File - assume the first row of the worksheet is the header.
        elif start_header_value == row[start_header_index].value :
            have_header_row = True
        if have_header_row :
            findRoleHeaders(row)
            min_data_row = row_ix + 1

if not have_header_row :
    MessageShow(f"Unable to locate the header row in \"{worksheet_name}\" worksheet.  Searched first 12 rows for \"{start_header_value}\" in column {start_header_index}.")
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
#
#   Verify all needed columns are present in Source and Target worksheets.  Build "move_list" list. (See "move_list.append(...)" below, for content.)
missing_required_header = False
key_is_int = False
move_list = []
for key, value in ctrl_dict.items():
    source_ix = -1
    source_column = value[0].value
    target_ix = -1
    target_column = key.value
    if  source_column in role_headers :
        source_ix = role_headers[source_column]
    else :
        MessageShow(f"Did not find required column header \"{source_column}\" on {worksheet_name} worksheet.  (Needed as source for \"{target_column}\" column.)")
        missing_required_header = True
    if  target_column in action_headers :
        target_ix = action_headers[target_column]
    else:
        MessageShow(f"Did not find required column header \"{target_column}\" on {actsheet_name} worksheet.  (Source from \"{source_column}\" column.)")
        missing_required_header = True
        
    move_list.append((source_ix, target_ix, value[1].value, value[2].value))
    
#   Identify key column in both source and target worksheets.  The 'key_column' parameter identifies the target (Action) worksheet header.
    if target_column == parms_dict['key_column'][0] :
        source_key_ix = source_ix
        target_key_ix = target_ix
        if value[1].value == "int" :
            key_is_int = True
        
if missing_required_header :
    MessageShow("Terminating process.")
    MessageClose()
    sys.exit(1)
    
   
#   Build dictionary of Request IDs in current Action Worksheet.   This is needed for comparing - if the same Request ID shows up in the Source Worksheet.
#
#   Note: target_key_ix was derived earlier (when building move_list), from the 'key_column' parameter.
action_requests = {} 
for act_row in actsheet.iter_rows(min_row=2, max_row=actsheet.max_row, values_only=False) :
    action_requests[act_row[target_key_ix].value] = act_row[0].row
    
    
#
#   Pull in user input for default action and date.  This is used for roles (rows) that have no action/date specified in the source worksheet.

default_action = input("What should be the default action value? ")
if default_action == "" :
    MessageShow("Ending process.")
    MessageClose()
    sys.exit(1)

have_date = False
while not have_date :
    default_date_str = input("What should be the default action date (mm/dd/year form) ? ").strip()
    if default_date_str == "" :
        MessageShow("Ending process.")
        MessageClose()
        sys.exit(1)
    try:
        default_date = datetime.strptime(default_date_str, "%m/%d/%Y")
        have_date = True
    except ValueError as e:
        print(f"Error parsing date: {e}")

#
#   splitCell() - Source cell consists of action (string) and date in mm/dd/year form.  (Delimited by space.)  Split into action and date, and pass back one or the other.
def splitCell(indicator, cell_content) :
    global default_action, default_date, key_value_display
    
    if cell_content == None :
        if indicator == "str" :
            return default_action
        else :
            return default_date

    cell_content = cell_content.strip()
    split_list = cell_content.split()

    split_len = len(split_list)
    if split_len == 0 :
        if indicator == "str" :
            return default_action
        else :
            return default_date
    if split_len == 1 :     # Unexpected 
        MessageShow("Unexpected Action/Date value, \"{cell_content}\" at {key_value_display}")
        if indicator == "str" :
            return cell_content
        else :
            return default_date

    
#   Verify that the last part of the cell content is a valid date in mm/dd/year form.
    date_str = split_list[-1]
    have_date = False
    try:
        action_date = datetime.strptime(date_str, "%m/%d/%Y")
        have_date = True
    except ValueError as e:
        MessageShow("Unexpected Action/Date value, \"{cell_content}\" at {key_value_display}")
        if indicator == "str" :
            return cell_content
        else :
            return default_date
           
    if indicator == "date" :
        return action_date
    else :
        date_ix = cell_content.find(split_list[-1])
        return cell_content[0:date_ix].strip()
 
    
show_process_warning = True
def buildRow(w_row) :
    global move_list, show_process_warning, key_value_display
    
    new_row = [""] * (max_action_col_ix + 1)
    warning_shown = False 
    for move_it in move_list :
        if move_it[3].strip() == "copy" :
            copy_value = w_row[move_it[0]].value
            if move_it[2].strip() == "int" :
                try :
                    copy_value = int(copy_value)
                except ValueError :
                    MessageShow(f"Unable to convert \"{copy_value}\" to int object at {key_value_display}")
                except TypeError :
                    MessageShow(f"Unable to convert \"{copy_value}\" to int object at {key_value_display}")
            new_row[move_it[1]] = copy_value
        elif move_it[3].strip() == "split" :
            new_row[move_it[1]] = splitCell(move_it[2], w_row[move_it[0]].value)
        elif show_process_warning :
            warning_shown = True
            MessageShow("Unable to process \"{move_it[3].strip()}\" indicator.")
    
    if warning_shown :      # Show warning for any bad process code - but only on the first call to buildRow.
        show_process_warning = False
        
    return new_row

action_updated = False
count_rows  = 1
count_match = 0
count_added = 0
count_type_error = 0
for w_row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=False) :
    count_rows += 1

    key_value = w_row[source_key_ix].value      # source_key_ix was derived earlier (when building move_list), from the 'key_column' parameter.
    key_value_display = f"Request # {key_value} at row {count_rows}"
    have_key_error = False
    if key_is_int :
        try :
            key_value = int(key_value)
        except ValueError :
            MessageShow(f"Unable to convert key value, \"{key_value}\" to int object, at {key_value_display}")
            have_key_error = True
        except TypeError :
            MessageShow(f"Unable to convert key value, \"{key_value}\" to int object, at {key_value_display}")
            have_key_error = True
    
    if have_key_error :
        count_type_error += 1
        if count_type_error > 11 :
            MessageShow(f"Ending process b/c key errors.  Worksheet size is {worksheet.max_row - 1}.")
            break
    elif key_value in action_requests :
        count_match += 1
    else :
        new_row = buildRow(w_row)
        if isinstance(new_row, list) :
            actsheet.append(new_row)
            action_updated = True
            count_added += 1
            for move_it in move_list :
                if move_it[2].strip() == "date" :
                    date_cell = actsheet.cell(row=actsheet.max_row, column=move_it[1] + 1)
                    date_cell.number_format = 'mm/dd/yyyy'

MessageShow(f"{count_rows} Requests read from {worksheet_name}.")
MessageShow(f"{count_added} Requests added to {actsheet_name}. {count_match} matching - already listed.")
if action_updated :
    input_dummy = input(f"Is the {actsheet_name} workbook closed?  If not, will CRASH !!!  <Enter> to continue.")
    actbook.save(actionfile_path)


MessageClose()